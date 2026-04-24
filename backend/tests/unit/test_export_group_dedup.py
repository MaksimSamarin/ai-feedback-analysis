"""Юнит-тесты на дедупликацию групп в xlsx-выгрузке (v2.0.0, итерация 3.2).

При групповом анализе LLM возвращает один ответ на группу, который применяется
ко всем строкам группы. В итоговый xlsx раньше попадали ВСЕ строки группы
(дубль: 1 группа = 100 строк с одинаковым analysis). Теперь в выгрузку идёт
одна строка на group_key (как и в preview).

Запуск:
    cd backend && pytest tests/unit/test_export_group_dedup.py -v
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _read_results_sheet(path: Path) -> list[list]:
    wb = load_workbook(path, read_only=False)
    ws = wb["results"]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _done_row(row_number: int, group_key: str | None, summary_text: str) -> dict:
    return {
        "row_number": row_number,
        "status": "done",
        "group_key": group_key,
        "custom_json": f'{{"summary": "{summary_text}"}}',
        "passthrough_json": None,
        "warnings": [],
        "error": None,
    }


def test_grouped_export_emits_one_row_per_group(tmp_path: Path) -> None:
    """Для группового отчёта в xlsx попадает одна строка на group_key."""
    from app.services.excel_service import export_results_xlsx

    rows = [
        _done_row(1, "store-A", "первая группа"),
        _done_row(2, "store-A", "первая группа"),
        _done_row(3, "store-A", "первая группа"),
        _done_row(4, "store-B", "вторая группа"),
        _done_row(5, "store-B", "вторая группа"),
        _done_row(6, "store-C", "третья группа"),
    ]

    def factory():
        return iter(rows)

    out = tmp_path / "grouped.xlsx"
    export_results_xlsx(out, factory, summary={"total_rows": 6})

    sheet = _read_results_sheet(out)
    # Первая строка — заголовки. Дальше — по одной на group_key.
    data_rows = sheet[1:]
    assert len(data_rows) == 3, f"Ожидалось 3 строки (по одной на группу), получено {len(data_rows)}"


def test_non_grouped_export_keeps_all_materialized_rows(tmp_path: Path) -> None:
    """Для негруппового отчёта (group_key=None) все done-строки попадают в xlsx."""
    from app.services.excel_service import export_results_xlsx

    rows = [
        _done_row(1, None, "first"),
        _done_row(2, None, "second"),
        _done_row(3, None, "third"),
    ]

    def factory():
        return iter(rows)

    out = tmp_path / "flat.xlsx"
    export_results_xlsx(out, factory, summary={"total_rows": 3})

    sheet = _read_results_sheet(out)
    data_rows = sheet[1:]
    assert len(data_rows) == 3


def test_grouped_export_skips_pending_rows(tmp_path: Path) -> None:
    """Pending-строки не попадают в выгрузку даже в групповом режиме."""
    from app.services.excel_service import export_results_xlsx

    rows = [
        _done_row(1, "A", "first"),
        {
            "row_number": 2,
            "status": "pending",
            "group_key": "A",
            "custom_json": None,
            "passthrough_json": None,
            "warnings": [],
            "error": None,
        },
        _done_row(3, "B", "second"),
    ]

    def factory():
        return iter(rows)

    out = tmp_path / "partial.xlsx"
    export_results_xlsx(out, factory, summary={"total_rows": 3})

    sheet = _read_results_sheet(out)
    data_rows = sheet[1:]
    # Группа A — одна done-строка, pending-строка группы A пропущена; группа B — одна строка.
    assert len(data_rows) == 2


def test_grouped_export_empty_group_key_treated_as_ungrouped(tmp_path: Path) -> None:
    """Строка с group_key='' (пусто) не дедуплицируется — ведёт себя как негрупповая."""
    from app.services.excel_service import export_results_xlsx

    rows = [
        _done_row(1, "", "row1"),
        _done_row(2, "", "row2"),
    ]

    def factory():
        return iter(rows)

    out = tmp_path / "empty_key.xlsx"
    export_results_xlsx(out, factory, summary={"total_rows": 2})

    sheet = _read_results_sheet(out)
    data_rows = sheet[1:]
    assert len(data_rows) == 2, "Пустой group_key — не дедуплицируем, ожидаем обе строки"


def test_grouped_export_keeps_only_group_column_from_passthrough(tmp_path: Path) -> None:
    """В групповом режиме в xlsx попадает только колонка группировки + агрегат LLM.

    Регрессия: раньше в xlsx для сгруппированного отчёта шли поля одной случайной
    строки группы (отзыв/оценка/дата одного клиента из смены) рядом с агрегатом
    по смене — это вводило в заблуждение.
    """
    import json as _json
    from openpyxl import load_workbook

    from app.services.excel_service import export_results_xlsx

    rows = [
        {
            "row_number": 1,
            "status": "done",
            "group_key": "Север_2026-03-01_утро",
            "custom_json": _json.dumps(
                {"число_нарушений": 3, "уровень_риска": "средний"},
                ensure_ascii=False,
            ),
            "passthrough_json": _json.dumps(
                {
                    "смена": "Север_2026-03-01_утро",
                    "оценка": 5,
                    "дата": "2026-03-01",
                },
                ensure_ascii=False,
            ),
            "warnings": [],
            "error": None,
        },
        {
            "row_number": 2,
            "status": "done",
            "group_key": "Север_2026-03-01_утро",
            "custom_json": _json.dumps(
                {"число_нарушений": 3, "уровень_риска": "средний"},
                ensure_ascii=False,
            ),
            "passthrough_json": _json.dumps(
                {"смена": "Север_2026-03-01_утро", "оценка": 1, "дата": "2026-03-01"},
                ensure_ascii=False,
            ),
            "warnings": [],
            "error": None,
        },
    ]

    def factory():
        return iter(rows)

    out = tmp_path / "grouped_cols.xlsx"
    export_results_xlsx(
        out, factory, summary={"total_rows": 2}, group_by_column="смена"
    )

    wb = load_workbook(out, read_only=False)
    ws = wb["results"]
    header = [c.value for c in ws[1]]
    assert "смена" in header
    assert "оценка" not in header, "В сгруппированной выгрузке не должно быть поля одной записи"
    assert "дата" not in header
    assert "число_нарушений" in header
    assert "уровень_риска" in header


def test_grouped_export_first_column_is_sequential_group_number(tmp_path: Path) -> None:
    """В групповой xlsx-выгрузке первая колонка — порядковый номер группы (1..N),
    а не row_number случайной строки из исходного файла.
    """
    from openpyxl import load_workbook

    from app.services.excel_service import export_results_xlsx

    rows = [
        _done_row(3107, "shift-A", "one"),
        _done_row(3108, "shift-A", "one"),
        _done_row(5589, "shift-B", "two"),
        _done_row(7041, "shift-C", "three"),
    ]

    def factory():
        return iter(rows)

    out = tmp_path / "group_ids.xlsx"
    export_results_xlsx(
        out, factory, summary={"total_rows": 4}, group_by_column="смена"
    )

    wb = load_workbook(out, read_only=False)
    ws = wb["results"]
    header = [c.value for c in ws[1]]
    assert header[0] == "group_number", "В заголовке xlsx для группового режима ожидаем 'group_number'"

    first_col = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert first_col == [1, 2, 3], f"Ожидали порядковые номера 1,2,3 — получили {first_col}"


def test_non_grouped_export_first_column_is_row_number(tmp_path: Path) -> None:
    """В негрупповой выгрузке первая колонка остаётся row_number (номер в исходном xlsx)."""
    from openpyxl import load_workbook

    from app.services.excel_service import export_results_xlsx

    rows = [
        _done_row(17, None, "first"),
        _done_row(42, None, "second"),
    ]

    def factory():
        return iter(rows)

    out = tmp_path / "flat_ids.xlsx"
    export_results_xlsx(out, factory, summary={"total_rows": 2})

    wb = load_workbook(out, read_only=False)
    ws = wb["results"]
    header = [c.value for c in ws[1]]
    assert header[0] == "row_number"
    first_col = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert first_col == [17, 42]
