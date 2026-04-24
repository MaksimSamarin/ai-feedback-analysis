"""Юнит-тест на стриминг-экспорт больших отчётов (BUG-14).

Проверяет что `export_results_xlsx`:
1. Не загружает все строки в память одновременно (работает с итератором, write-only workbook)
2. Корректно собирает заголовки из всех ключей при двух проходах
3. Пишет все строки в файл

Запуск:
    cd backend && pytest tests/unit/test_export_streaming.py -v
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterator

import pytest
from openpyxl import load_workbook

from app.services.excel_service import export_results_xlsx


def _make_row(idx: int) -> dict[str, Any]:
    """Генерит фейковую строку отчёта в формате, который возвращает `iter_report_rows`."""
    return {
        "row_number": idx,
        "review_text": f"Отзыв номер {idx}",
        "input_json": json.dumps({"text": f"input {idx}"}, ensure_ascii=False),
        "passthrough_json": json.dumps({"author": f"User {idx}"}, ensure_ascii=False),
        "analysis_json": {
            "sentiment_label": "negative" if idx % 3 == 0 else "positive",
            "negativity_score": 0.5,
            "summary": f"Итог строки {idx}",
            "category": "service",
        },
        "warnings": [],
        "error": None,
    }


def test_export_results_xlsx_streams_without_materializing_rows(tmp_path: Path) -> None:
    """Экспорт не должен превращать итератор в список в памяти.

    Признак стриминга: рабочая книга пишется в режиме `write_only=True`,
    фабрика итераторов вызывается ровно два раза (пас сбора заголовков + пас записи),
    сам итератор — генератор (а не список), и все строки доходят до файла.
    """
    total_rows = 5_000
    factory_call_count = 0
    yielded_rows_pass: list[int] = []

    def rows_factory() -> Iterator[dict[str, Any]]:
        nonlocal factory_call_count
        factory_call_count += 1

        def gen() -> Iterator[dict[str, Any]]:
            produced = 0
            for i in range(1, total_rows + 1):
                produced += 1
                yield _make_row(i)
            yielded_rows_pass.append(produced)

        return gen()

    out_path = tmp_path / "out.xlsx"
    export_results_xlsx(
        out_path,
        rows_factory,
        summary={"total_rows": total_rows, "success_rows": total_rows},
        prompt_example="Проанализируй отзыв: {row_json}",
    )

    # 1. Factory вызвана ровно дважды: сначала для сбора заголовков, потом для записи
    assert factory_call_count == 2, (
        f"ожидалось 2 вызова фабрики (discovery + write), получено {factory_call_count}"
    )

    # 2. Обе итерации прошли полностью
    assert yielded_rows_pass == [total_rows, total_rows], (
        f"итератор обошёл не все строки: {yielded_rows_pass}"
    )

    # 3. Файл создан
    assert out_path.exists(), "xlsx-файл не создан"

    # 4. Листы на месте и количество строк соответствует входу
    wb = load_workbook(out_path, read_only=True)
    try:
        assert "results" in wb.sheetnames
        assert "summary" in wb.sheetnames
        assert "prompt_example" in wb.sheetnames

        ws = wb["results"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        assert header_row[0] == "row_number"
        assert "warnings" in header_row
        assert "error" in header_row
        # review_text — служебный внутренний текст для модели, в итоговый xlsx не пишется.
        assert "review_text" not in header_row, "review_text не должен попадать в итоговый xlsx"
        # Ключи из passthrough_json / analysis_json должны попасть в заголовок.
        # Ключи из input_json — НЕ должны: исходные колонки анализа ушли в LLM, в итоговом файле
        # остаются только справочные поля (passthrough) и результат модели.
        assert "text" not in header_row, "ключ из input_json не должен попадать в итоговый xlsx"
        assert "author" in header_row, "ключ из passthrough_json не добавлен в заголовок"
        assert "sentiment_label" in header_row, "ключ из analysis_json не добавлен в заголовок"

        data_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        assert data_rows == total_rows, (
            f"в файле {data_rows} строк данных, ожидалось {total_rows}"
        )
    finally:
        wb.close()


def test_export_results_xlsx_uses_write_only_mode(tmp_path: Path) -> None:
    """Регрессия: экспорт должен использовать Workbook(write_only=True).

    Workbook без write_only держит все строки в памяти до save() — на больших отчётах это
    приводило к OOM-kill (BUG-14). Проверяем через инстроспекцию: у write-only книги
    нет активного листа (`wb.active is None`), и приходится создавать через `create_sheet`.
    """
    import openpyxl

    original_workbook = openpyxl.Workbook
    observed_kwargs: list[dict[str, Any]] = []

    def tracking_workbook(*args: Any, **kwargs: Any) -> Any:
        observed_kwargs.append(kwargs)
        return original_workbook(*args, **kwargs)

    monkey_target = "app.services.excel_service.Workbook"
    # Патчим через monkeypatch напрямую — чтобы не зависеть от fixture.
    import app.services.excel_service as mod

    mod.Workbook = tracking_workbook  # type: ignore[assignment]
    try:
        def factory() -> Iterator[dict[str, Any]]:
            yield _make_row(1)

        export_results_xlsx(
            tmp_path / "one.xlsx",
            factory,
            summary={"total_rows": 1},
        )
    finally:
        mod.Workbook = original_workbook  # type: ignore[assignment]

    assert observed_kwargs, "Workbook не был вызван"
    # Хотя бы один вызов должен быть с write_only=True (основной экспорт)
    assert any(kw.get("write_only") is True for kw in observed_kwargs), (
        f"Workbook должен инстанцироваться с write_only=True, kwargs={observed_kwargs}"
    )


def test_export_results_xlsx_excludes_analysis_input_columns(tmp_path: Path) -> None:
    """В итоговом xlsx исходные колонки анализа не пишутся отдельными столбцами.

    Бизнес-требование v2.0.0: в итог идут только справочные колонки (passthrough)
    и результат модели. Колонки, ушедшие в LLM, уже отражены в `review_text`.
    """
    def rows_factory() -> Iterator[dict[str, Any]]:
        def gen() -> Iterator[dict[str, Any]]:
            yield {
                "row_number": 1,
                "review_text": "review: Очень плохо\nstars: 1",
                "input_json": json.dumps({"review": "Очень плохо", "stars": "1"}, ensure_ascii=False),
                "passthrough_json": json.dumps({"order_id": "A-42"}, ensure_ascii=False),
                "analysis_json": {"verdict": "негатив"},
                "warnings": [],
                "error": None,
            }

        return gen()

    out_path = tmp_path / "no_input.xlsx"
    export_results_xlsx(
        out_path,
        rows_factory,
        summary={"total_rows": 1},
    )

    wb = load_workbook(out_path, read_only=True)
    try:
        ws = wb["results"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        # Исходные колонки анализа в итог не попали:
        assert "review" not in header_row
        assert "stars" not in header_row
        # review_text — служебный для модели, в итог не пишется:
        assert "review_text" not in header_row
        # Справочная колонка попала:
        assert "order_id" in header_row
        # Результат модели попал:
        assert "verdict" in header_row
        # Служебные на месте:
        assert "warnings" in header_row
        assert "error" in header_row

        data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        data_map = dict(zip(header_row, data_row))
        assert data_map["order_id"] == "A-42"
        assert data_map["verdict"] == "негатив"
    finally:
        wb.close()


def test_export_results_xlsx_reads_llm_result_from_custom_json(tmp_path: Path) -> None:
    """Регрессия: LLM-результат хранится в report_rows.custom_json (TEXT, JSON-строка),
    а не в `analysis_json`. В предыдущей реализации `export_results_xlsx` читал только
    `analysis_json` — из-за этого итоговый xlsx оставался без колонок анализа, хотя
    модель отвечала корректно. Проверяем что _extract_analysis корректно парсит
    строку custom_json и разворачивает её в колонки результата.
    """
    def rows_factory() -> Iterator[dict[str, Any]]:
        def gen() -> Iterator[dict[str, Any]]:
            # Ровно то, что возвращает iter_report_rows из Postgres — custom_json
            # как JSON-строка, analysis_json в row отсутствует.
            yield {
                "row_number": 1,
                "review_text": "review: Хорошо",
                "input_json": json.dumps({"review": "Хорошо"}, ensure_ascii=False),
                "passthrough_json": json.dumps({"order_id": "B-1"}, ensure_ascii=False),
                "custom_json": json.dumps(
                    {"verdict": "позитив", "score": 0.85},
                    ensure_ascii=False,
                ),
                "warnings": [],
                "error": None,
            }

        return gen()

    out_path = tmp_path / "from_custom_json.xlsx"
    export_results_xlsx(out_path, rows_factory, summary={"total_rows": 1})

    wb = load_workbook(out_path, read_only=True)
    try:
        ws = wb["results"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        # Ключи из custom_json ДОЛЖНЫ попасть в заголовок — это и было сломано
        assert "verdict" in header, "LLM-ключ из custom_json не попал в xlsx (регресс)"
        assert "score" in header
        assert "order_id" in header
        data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        data = dict(zip(header, data_row))
        assert data["verdict"] == "позитив"
        assert data["score"] == 0.85
        assert data["order_id"] == "B-1"
    finally:
        wb.close()


def test_export_results_xlsx_allows_column_both_in_analysis_and_passthrough(tmp_path: Path) -> None:
    """Пользователь может выбрать одну и ту же колонку и для анализа, и в итоговый отчёт.

    В этом случае на уровне БД значение дублируется в input_json и passthrough_json;
    в xlsx оно попадает один раз — как passthrough (т.к. input_json в xlsx не пишется).
    """
    def rows_factory() -> Iterator[dict[str, Any]]:
        def gen() -> Iterator[dict[str, Any]]:
            yield {
                "row_number": 1,
                "review_text": "comment: Хорошо",
                "input_json": json.dumps({"comment": "Хорошо"}, ensure_ascii=False),
                "passthrough_json": json.dumps({"comment": "Хорошо"}, ensure_ascii=False),
                "analysis_json": {"verdict": "позитив"},
                "warnings": [],
                "error": None,
            }

        return gen()

    out_path = tmp_path / "overlap.xlsx"
    export_results_xlsx(
        out_path,
        rows_factory,
        summary={"total_rows": 1},
    )

    wb = load_workbook(out_path, read_only=True)
    try:
        ws = wb["results"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        assert header_row.count("comment") == 1
        data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        data_map = dict(zip(header_row, data_row))
        assert data_map["comment"] == "Хорошо"
        assert data_map["verdict"] == "позитив"
    finally:
        wb.close()


@pytest.mark.slow
def test_export_results_xlsx_memory_stays_bounded_on_large_report(tmp_path: Path) -> None:
    """Регрессия на BUG-14: на отчёте сопоставимом с реальным (сотни тысяч строк)
    пик RSS не должен расти линейно с числом строк.

    Пользователь пожаловался что падает на ~1 млн строк. 200k здесь хватает чтобы
    отличить стриминг (O(1) по памяти) от старой реализации (O(N)): старая съела бы
    200–400 МБ на такой объём, стриминг укладывается в ≪100 МБ прироста RSS.
    """
    import gc
    import psutil

    total_rows = 200_000

    def rows_factory() -> Iterator[dict[str, Any]]:
        def gen() -> Iterator[dict[str, Any]]:
            for i in range(1, total_rows + 1):
                yield _make_row(i)

        return gen()

    out_path = tmp_path / "large.xlsx"
    process = psutil.Process()

    gc.collect()
    rss_before_mb = process.memory_info().rss / 1024 / 1024

    export_results_xlsx(
        out_path,
        rows_factory,
        summary={"total_rows": total_rows, "success_rows": total_rows},
    )

    rss_after_mb = process.memory_info().rss / 1024 / 1024
    rss_delta_mb = rss_after_mb - rss_before_mb

    # Порог 200 МБ с запасом: openpyxl write_only сам по себе берёт память под
    # буферы и shared strings. На старой реализации на таком объёме было бы
    # >500 МБ; на стриминге — 20–100 МБ на практике.
    assert rss_delta_mb < 200, (
        f"BUG-14 регресс: пик RSS вырос на {rss_delta_mb:.0f} МБ "
        f"при {total_rows} строк, ожидалось < 200 МБ (признак стриминга)"
    )

    assert out_path.exists()
    assert out_path.stat().st_size > 0
