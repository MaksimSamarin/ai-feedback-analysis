"""Юнит-тест на сборку xlsx из частичного состояния БД (итерация 2, волна E).

Когда пользователь запрашивает `/download/partial/xlsx` на running/paused-отчёте —
backend должен собрать файл из строк, которые **уже обработаны** (done/error),
а pending-строки пропустить. Это критично для больших отчётов: на 1 млн строк
при прогрессе 10k не таскать в xlsx 990k пустых записей.

Проверяем через мок-стаб `iter_report_rows`: 3 строки — одна готова, одна
с ошибкой, одна ещё pending. В итоговом xlsx должно быть ровно 2 строки
(1 и 2), row_number=3 отсутствует.

Запуск:
    cd backend && pytest tests/unit/test_partial_export_from_db.py -v
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from app.services.excel_service import export_results_xlsx


def _make_partial_rows() -> list[dict[str, Any]]:
    """3 строки: готовая, с ошибкой, pending (с явным status)."""
    return [
        {
            "row_number": 1,
            "status": "done",
            "review_text": "review: Очень хорошо",
            "input_json": json.dumps({"review": "Очень хорошо"}, ensure_ascii=False),
            "passthrough_json": json.dumps({"order_id": "A-1"}, ensure_ascii=False),
            "analysis_json": {"verdict": "позитив"},
            "warnings": [],
            "error": None,
        },
        {
            "row_number": 2,
            "status": "error",
            "review_text": "review: Ужасно",
            "input_json": json.dumps({"review": "Ужасно"}, ensure_ascii=False),
            "passthrough_json": json.dumps({"order_id": "A-2"}, ensure_ascii=False),
            "analysis_json": None,
            "warnings": ["invalid_json"],
            "error": "Модель вернула невалидный JSON",
        },
        {
            "row_number": 3,
            "status": "pending",
            "review_text": "review: Нормально",
            "input_json": json.dumps({"review": "Нормально"}, ensure_ascii=False),
            "passthrough_json": json.dumps({"order_id": "A-3"}, ensure_ascii=False),
            "analysis_json": None,
            "warnings": [],
            "error": None,
        },
    ]


def test_partial_export_skips_pending_rows(tmp_path: Path) -> None:
    """Pending-строки не попадают в xlsx — иначе на миллионных отчётах тянем 90% мусора.

    Должны попасть только строки с status in ('done', 'error'). review_text больше
    тоже не пишется как отдельная колонка (он был служебным для модели).
    """
    rows = _make_partial_rows()

    def factory() -> Iterator[dict[str, Any]]:
        for row in rows:
            yield row

    out_path = tmp_path / "partial.xlsx"
    export_results_xlsx(
        out_path,
        factory,
        summary={
            "total_rows": 3,
            "success_rows": 1,
            "failed_rows": 1,
            "partial": True,
        },
    )

    wb = load_workbook(out_path, read_only=True)
    try:
        ws = wb["results"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        assert "row_number" in header
        assert "order_id" in header
        assert "verdict" in header
        assert "warnings" in header
        assert "error" in header
        # review_text — служебный, не тянется в итог
        assert "review_text" not in header

        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Только 2 обработанные строки, pending отфильтрована:
        assert len(raw_rows) == 2, f"ожидалось 2 обработанные строки, получили {len(raw_rows)}"

        def _padded(row: tuple) -> tuple:
            if len(row) >= len(header):
                return row
            return row + (None,) * (len(header) - len(row))
        data_rows = [_padded(r) for r in raw_rows]

        row_by_num = {row[header.index("row_number")]: row for row in data_rows}
        assert set(row_by_num.keys()) == {1, 2}, "pending строка 3 не должна попасть в xlsx"

        # Первая — результат есть
        first = row_by_num[1]
        assert first[header.index("verdict")] == "позитив"
        assert first[header.index("order_id")] == "A-1"

        # Вторая — error заполнен, verdict пуст
        second = row_by_num[2]
        assert second[header.index("verdict")] is None
        assert second[header.index("error")] == "Модель вернула невалидный JSON"

        # Summary содержит признак partial и оригинальный total_rows=3
        # (включая pending — это общий размер файла, а не размер выгрузки).
        summary_ws = wb["summary"]
        summary_pairs = {str(row[0]): row[1] for row in summary_ws.iter_rows(min_row=2, values_only=True)}
        assert summary_pairs.get("partial") in (True, "True", "true")
        assert summary_pairs.get("total_rows") == 3
    finally:
        wb.close()


def test_partial_export_falls_back_on_data_when_status_missing(tmp_path: Path) -> None:
    """Устойчивость к строкам без явного status: если analysis_json/error_text есть —
    строка считается обработанной. Это бэк-совместимость с тестами/моками без status."""
    def factory() -> Iterator[dict[str, Any]]:
        yield {
            "row_number": 1,
            "review_text": "x",
            "passthrough_json": json.dumps({"author": "a"}, ensure_ascii=False),
            "analysis_json": {"verdict": "ок"},
            "warnings": [],
            "error": None,
        }
        # Нет status, нет analysis_json, нет error — считается pending и отфильтровывается
        yield {
            "row_number": 2,
            "review_text": "y",
            "passthrough_json": json.dumps({"author": "b"}, ensure_ascii=False),
            "analysis_json": None,
            "warnings": [],
            "error": None,
        }

    out_path = tmp_path / "fallback.xlsx"
    export_results_xlsx(out_path, factory, summary={"total_rows": 2})

    wb = load_workbook(out_path, read_only=True)
    try:
        ws = wb["results"]
        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(raw_rows) == 1  # Только первая материализованная попала
    finally:
        wb.close()
