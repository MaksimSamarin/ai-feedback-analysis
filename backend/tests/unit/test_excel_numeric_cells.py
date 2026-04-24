"""Юнит-тесты на сохранение числовых типов при экспорте xlsx.

Итерация 1 отказа от обязательных полей предполагает что пользователь сам
определяет колонки вывода, включая числовые (оценка, срочность, confidence).
Нужно убедиться что `int`/`float`/`bool` из `custom_json` записываются в xlsx
как number/boolean, а не строка — иначе формулы Excel не работают.

Проверяем через запись в BytesIO и чтение обратно — тип ячейки должен быть
Number или Boolean, не TEXT.

Запуск:
    cd backend && pytest tests/unit/test_excel_numeric_cells.py -v
"""

from __future__ import annotations

import io
import json


def _export_and_read(rows: list[dict]):
    """Экспортирует через export_results_xlsx в BytesIO и читает обратно openpyxl'ом."""
    from app.services.excel_service import export_results_xlsx
    from openpyxl import load_workbook

    def rows_factory():
        for row in rows:
            yield row

    buf = io.BytesIO()
    export_results_xlsx(
        buf,
        rows_factory,
        {"total_rows": len(rows), "processed_rows": len(rows), "success_rows": len(rows), "failed_rows": 0},
    )
    buf.seek(0)
    wb = load_workbook(buf)
    return wb


def test_int_stays_number_in_xlsx() -> None:
    """int из custom_json → ячейка с типом Number, а не строка."""
    wb = _export_and_read(
        [
            {
                "row_number": 1,
                "review_text": "опоздание",
                "input_json": "{}",
                "passthrough_json": "{}",
                "analysis_json": {"срочность": 9},
                "warnings": [],
                "error": None,
            }
        ]
    )
    ws = wb.active

    # Найти колонку "срочность"
    headers = [cell.value for cell in ws[1]]
    assert "срочность" in headers
    col_idx = headers.index("срочность") + 1  # 1-based

    cell = ws.cell(row=2, column=col_idx)
    assert cell.value == 9
    assert isinstance(cell.value, int)


def test_float_stays_number_in_xlsx() -> None:
    """float → number, не теряет знак после запятой."""
    wb = _export_and_read(
        [
            {
                "row_number": 1,
                "review_text": "x",
                "input_json": "{}",
                "passthrough_json": "{}",
                "analysis_json": {"рейтинг": 4.75},
                "warnings": [],
                "error": None,
            }
        ]
    )
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_idx = headers.index("рейтинг") + 1
    cell = ws.cell(row=2, column=col_idx)
    assert cell.value == 4.75
    assert isinstance(cell.value, float)


def test_bool_stays_bool_in_xlsx() -> None:
    """bool → сохраняется как TRUE/FALSE, openpyxl отдаёт bool."""
    wb = _export_and_read(
        [
            {
                "row_number": 1,
                "review_text": "x",
                "input_json": "{}",
                "passthrough_json": "{}",
                "analysis_json": {"требует_эскалации": True},
                "warnings": [],
                "error": None,
            }
        ]
    )
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_idx = headers.index("требует_эскалации") + 1
    cell = ws.cell(row=2, column=col_idx)
    assert cell.value is True


def test_int_in_passthrough_stays_number() -> None:
    """Числа в passthrough (из исходного Excel) тоже остаются числами."""
    wb = _export_and_read(
        [
            {
                "row_number": 1,
                "review_text": "x",
                "input_json": "{}",
                "passthrough_json": json.dumps({"price": 1250}),
                "analysis_json": {},
                "warnings": [],
                "error": None,
            }
        ]
    )
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_idx = headers.index("price") + 1
    cell = ws.cell(row=2, column=col_idx)
    assert cell.value == 1250
    assert isinstance(cell.value, int)


def test_mixed_numeric_types_in_same_row() -> None:
    """Смесь int/float/bool/str в одной строке — каждый тип сохраняет свою сущность."""
    wb = _export_and_read(
        [
            {
                "row_number": 1,
                "review_text": "x",
                "input_json": "{}",
                "passthrough_json": "{}",
                "analysis_json": {
                    "категория": "доставка",
                    "score_int": 42,
                    "score_float": 3.14,
                    "critical": False,
                },
                "warnings": [],
                "error": None,
            }
        ]
    )
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    def _cell(name: str):
        idx = headers.index(name) + 1
        return ws.cell(row=2, column=idx).value

    assert _cell("категория") == "доставка"
    assert _cell("score_int") == 42 and isinstance(_cell("score_int"), int)
    assert _cell("score_float") == 3.14 and isinstance(_cell("score_float"), float)
    assert _cell("critical") is False
