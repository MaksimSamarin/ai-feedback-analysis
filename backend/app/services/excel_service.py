from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator

from openpyxl import Workbook
from python_calamine import CalamineWorkbook


RowsFactory = Callable[[], Iterable[dict[str, Any]]]


EMPTY_MARKERS = {"", "nan", "none", "null", "n/a", "na", "-"}


# Верхняя граница для подсчёта уникальных значений на колонку во время inspect.
# Обоснование: inspect-worker имеет mem_limit=2g (docker-compose.yml).
# Один set в Python ~200 байт на элемент; при 20 колонках и cap=100_000
# занято ~400 MB — комфортно вписывается в бюджет.
# Покрывает реальные сценарии группировки (операторы, сессии, клиенты за квартал).
# Переопределяется ENV INSPECT_UNIQUE_CAP.
INSPECT_UNIQUE_CAP = int(os.getenv("INSPECT_UNIQUE_CAP", "100000"))


def inspect_xlsx(path: Path, *, unique_cap: int = INSPECT_UNIQUE_CAP) -> list[dict[str, Any]]:
    wb = CalamineWorkbook.from_path(str(path))
    sheets: list[dict[str, Any]] = []

    try:
        for sheet_name in wb.sheet_names:
            ws = wb.get_sheet_by_name(sheet_name)
            rows = ws.iter_rows()
            header = next(rows, [])
            columns = [str(cell).strip() if cell is not None else "" for cell in header]
            columns = [f"Column {idx + 1}" if not col else col for idx, col in enumerate(columns)]

            unique_buckets: list[set[Any] | None] = [set() for _ in columns]
            total_rows = 0
            for row in rows:
                total_rows += 1
                for idx, cell in enumerate(row):
                    if idx >= len(unique_buckets):
                        continue
                    bucket = unique_buckets[idx]
                    if bucket is None:
                        continue
                    if cell is None:
                        continue
                    if isinstance(cell, str):
                        key: Any = cell.strip()
                        if not key:
                            continue
                    elif isinstance(cell, (int, float, bool)):
                        key = cell
                    else:
                        key = str(cell)
                    bucket.add(key)
                    if len(bucket) > unique_cap:
                        unique_buckets[idx] = None

            unique_counts: dict[str, int | None] = {}
            for idx, col in enumerate(columns):
                bucket = unique_buckets[idx]
                unique_counts[col] = len(bucket) if bucket is not None else None

            sheets.append(
                {
                    "name": ws.name,
                    "columns": columns,
                    "total_rows": total_rows,
                    "unique_counts": unique_counts,
                }
            )
    finally:
        wb.close()

    return sheets


def normalize_review(value: Any) -> tuple[str | None, list[str]]:
    warnings: list[str] = []
    if value is None:
        return None, ["empty_cell"]

    text = str(value).strip()
    if text.lower() in EMPTY_MARKERS:
        return None, ["empty_cell"]

    return text, warnings


def _build_review_text(row_payload: dict[str, str | None]) -> str | None:
    parts = [f"{key}: {value}" for key, value in row_payload.items() if value not in (None, "")]
    if not parts:
        return None
    return "\n".join(parts)


def iter_sheet_rows(
    path: Path,
    *,
    sheet_name: str,
    analysis_columns: list[str],
    max_reviews: int,
    non_analysis_columns: list[str] | None = None,
) -> Iterator[dict[str, Any]]:
    wb = CalamineWorkbook.from_path(str(path))
    try:
        if sheet_name not in wb.sheet_names:
            raise ValueError(f"Лист не найден: {sheet_name}")

        ws = wb.get_sheet_by_name(sheet_name)
        rows = ws.iter_rows()
        header = next(rows, [])
        columns = [str(cell).strip() if cell is not None else "" for cell in header]
        columns = [f"Column {idx + 1}" if not col else col for idx, col in enumerate(columns)]

        selected_analysis_columns = [col for col in analysis_columns if col]
        selected_non_analysis_columns = [col for col in (non_analysis_columns or []) if col]
        if not selected_analysis_columns:
            raise ValueError("Нужно выбрать хотя бы одну колонку для анализа")
        for selected_column in [*selected_analysis_columns, *selected_non_analysis_columns]:
            if selected_column not in columns:
                raise ValueError(f"Колонка не найдена: {selected_column}")

        analysis_indices = {name: columns.index(name) for name in selected_analysis_columns}
        passthrough_indices = {name: columns.index(name) for name in selected_non_analysis_columns}

        yielded = 0
        for row_number, row in enumerate(rows, start=2):
            if yielded >= max_reviews:
                break

            warnings: list[str] = []
            row_payload: dict[str, str | None] = {}
            passthrough_payload: dict[str, str | None] = {}
            for col_name, idx in analysis_indices.items():
                value = row[idx] if idx < len(row) else None
                value_text, value_warnings = normalize_review(value)
                row_payload[col_name] = value_text
                for warning in value_warnings:
                    if warning not in warnings:
                        warnings.append(warning)
            for col_name, idx in passthrough_indices.items():
                value = row[idx] if idx < len(row) else None
                passthrough_payload[col_name] = None if value is None else str(value).strip()

            text = _build_review_text(row_payload)

            yielded += 1
            yield {
                "row_number": row_number,
                "review_text": text,
                "input_json": json.dumps(row_payload, ensure_ascii=False),
                "passthrough_json": json.dumps(passthrough_payload, ensure_ascii=False),
                "warnings": warnings,
            }
    finally:
        wb.close()


def read_sheet_rows(
    path: Path,
    *,
    sheet_name: str,
    analysis_columns: list[str],
    max_reviews: int,
    non_analysis_columns: list[str] | None = None,
) -> list[dict[str, Any]]:
    return list(
        iter_sheet_rows(
            path,
            sheet_name=sheet_name,
            analysis_columns=analysis_columns,
            max_reviews=max_reviews,
            non_analysis_columns=non_analysis_columns,
        )
    )


_RESERVED_INPUT_KEYS = frozenset({"row_number", "review_text", "input_json", "warnings", "error"})


def _flatten_dict(data: dict[str, Any], prefix: str = "") -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in data.items():
        col = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            out.update(_flatten_dict(value, col))
        else:
            out[col] = value
    return out


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item is not None)
    return value


def _parse_json_field(raw: Any) -> dict[str, Any]:
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw.strip():
        try:
            decoded = json.loads(raw)
            if isinstance(decoded, dict):
                return decoded
        except Exception:
            return {}
    return {}


def _row_is_materialized(row: dict[str, Any]) -> bool:
    """Строка считается «обработанной» если LLM дал результат или явно упала с ошибкой.

    Нужен для фильтрации pending-строк в выгрузке: в отчёте на 1 млн строк мы
    не должны тащить в xlsx 990k пустых записей, если обработано только 10k.
    Логика устойчива к обоим способам сигнализации:
    - status='done' или 'error' (если поле status передано)
    - иначе — наличие custom_json или error_text
    """
    status = row.get("status")
    if isinstance(status, str):
        if status in {"done", "error"}:
            return True
        if status in {"pending", "in_progress", "queued"}:
            return False
        # unknown status — fallback на данные
    if row.get("custom_json") is not None:
        return True
    if row.get("analysis_json") is not None:
        return True
    if row.get("error_text") or row.get("error"):
        return True
    return False


def _extract_analysis(row: dict[str, Any]) -> dict[str, Any]:
    """Возвращает разобранный JSON с результатом LLM-анализа строки.

    В БД поле называется `custom_json` (TEXT), но в тестах и устаревших местах
    иногда передают уже распарсенный dict под ключом `analysis_json`. Поддерживаем
    оба варианта ради обратной совместимости тестов.
    """
    custom = row.get("custom_json")
    if custom is not None:
        parsed = _parse_json_field(custom)
        if parsed:
            return parsed
    legacy = row.get("analysis_json")
    if isinstance(legacy, dict):
        return legacy
    return _parse_json_field(legacy)


def _collect_column_keys(
    rows: Iterable[dict[str, Any]],
    *,
    group_by_column: str | None = None,
) -> tuple[list[str], list[str]]:
    """Первый проход — собирает уникальные ключи passthrough/dynamic без хранения строк.

    Исходные колонки анализа (input_json) в итоговый xlsx не попадают:
    они уже отражены в `review_text` и дублировать их нет смысла. Если
    пользователю нужны справочные поля из исходника — он явно указывает
    их в `non_analysis_columns` (они идут в passthrough_json).

    Для группового отчёта (`group_by_column` задан) из passthrough остаётся
    только сама колонка группировки: остальные поля относятся к одной
    конкретной строке группы, а выгружается одна строка на группу.
    """
    passthrough_keys: list[str] = []
    seen_passthrough: set[str] = set()
    dynamic_keys: list[str] = []
    seen_dynamic: set[str] = set()

    for row in rows:
        if not _row_is_materialized(row):
            continue
        for key in _parse_json_field(row.get("passthrough_json")).keys():
            key_s = str(key)
            if key_s in _RESERVED_INPUT_KEYS or key_s in seen_passthrough:
                continue
            if group_by_column and key_s != group_by_column:
                continue
            seen_passthrough.add(key_s)
            passthrough_keys.append(key_s)
        analysis = _extract_analysis(row)
        for key in _flatten_dict(analysis or {}).keys():
            if key in seen_dynamic:
                continue
            seen_dynamic.add(key)
            dynamic_keys.append(key)

    return passthrough_keys, dynamic_keys


def export_results_xlsx(
    path: Path,
    rows_factory: RowsFactory,
    summary: dict[str, Any],
    *,
    prompt_example: str | None = None,
    group_by_column: str | None = None,
) -> None:
    """Стримит результаты отчёта в xlsx без загрузки всех строк в память.

    `rows_factory` — функция, возвращающая свежий итератор строк. Вызывается дважды:
    первый раз для сбора заголовков колонок, второй — для записи данных.
    Используется `Workbook(write_only=True)` — строки пишутся на диск по мере поступления.
    См. BUG-14.

    Для групповых отчётов (`group_by_column` задан) выгружается ОДНА строка на группу,
    и из passthrough остаётся только сама колонка группировки — остальные поля
    (отзыв, оценка, дата конкретной записи) относятся к одной случайной строке группы
    и в агрегате по группе смысла не несут.
    """
    passthrough_keys, dynamic_keys = _collect_column_keys(
        rows_factory(), group_by_column=group_by_column
    )

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("results")

    is_grouped = bool(group_by_column)
    # В групповом режиме в первой колонке идёт порядковый номер группы (1..N),
    # а не row_number одной случайной строки-представителя.
    first_column_header = "group_number" if is_grouped else "row_number"
    headers = [
        first_column_header,
        *passthrough_keys,
        *dynamic_keys,
        "warnings",
        "error",
    ]
    ws.append(headers)

    seen_group_keys: set[str] = set()
    group_counter = 0
    for row in rows_factory():
        # Пропускаем ещё не обработанные строки: они в БД с custom_json=NULL
        # и без них мы не «раздуваем» итоговый файл на миллионы пустых записей
        # (актуально для partial-выгрузки, но безопасно и для финализированной).
        if not _row_is_materialized(row):
            continue
        group_key = row.get("group_key")
        if group_key is not None and str(group_key) != "":
            gk = str(group_key)
            if gk in seen_group_keys:
                continue
            seen_group_keys.add(gk)
        parsed_passthrough = _parse_json_field(row.get("passthrough_json"))
        analysis = _extract_analysis(row)
        flat_analysis = _flatten_dict(analysis or {})

        if is_grouped:
            group_counter += 1
            first_column_value: Any = group_counter
        else:
            first_column_value = row.get("row_number")

        ws.append(
            [
                first_column_value,
                *[_normalize_cell(parsed_passthrough.get(key)) for key in passthrough_keys],
                *[_normalize_cell(flat_analysis.get(key)) for key in dynamic_keys],
                ", ".join(row.get("warnings") or []),
                row.get("error"),
            ]
        )

    summary_ws = wb.create_sheet("summary")
    summary_ws.append(["metric", "value"])
    for key, value in summary.items():
        summary_ws.append(
            [key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value]
        )

    if prompt_example:
        prompts_ws = wb.create_sheet("prompt_example")
        prompts_ws.append(["name", "value"])
        prompts_ws.append(["prompt", prompt_example])

    wb.save(path)


def export_raw_json(
    path: Path,
    *,
    rows_factory: RowsFactory,
    model: str,
    provider: str,
    prompt_template: str,
    app_version: str,
) -> None:
    """Стримит сырой JSON отчёта на диск без загрузки всех строк в память (см. BUG-14)."""
    metadata = {
        "provider": provider,
        "model": model,
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "prompt_template": prompt_template,
        "app_version": app_version,
    }
    with path.open("w", encoding="utf-8") as fh:
        fh.write("{\n  \"metadata\": ")
        json.dump(metadata, fh, ensure_ascii=False, indent=2)
        fh.write(",\n  \"rows\": [")
        first = True
        for row in rows_factory():
            row_copy = dict(row)
            row_copy.pop("input_json", None)
            if not first:
                fh.write(",")
            first = False
            fh.write("\n    ")
            json.dump(row_copy, fh, ensure_ascii=False)
        fh.write("\n  ]\n}\n")
