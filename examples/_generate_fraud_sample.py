"""Генератор тестового файла «Мошенничество_N_отзывов.xlsx».

100 000 отзывов клиентов по 5 магазинам. Русские имена колонок. Реалистичное
распределение: 94% обычных отзывов (сервис/брак/позитив/нейтрал), 6% — сознательно
подмешанные признаки мошенничества сотрудников (для демо и теста).

Магазины отличаются характером:
- Магазин_Север — «проблемный»: 15% fraud, вымогательство и оплата мимо кассы.
- Магазин_Запад — 10% fraud: курьеры и подмены товара.
- Магазин_Центр — 5% fraud: единичные случаи.
- Магазин_Юг — 0% fraud: сервисные жалобы без признаков мошенничества.
- Магазин_Онлайн — 0% fraud: технические проблемы (бот, приложение, доставка).

Колонка «смена» — ключ для РАБОЧЕЙ группировки (мелкие группы ~50 отзывов,
влезают в контекст LLM): «Север_2026-03-15_утро». При группировке по
«магазин» группы по 20k отзывов гарантированно не влезут в контекст — это
демонстрация защитной логики сервиса (honest error «Промпт превысил контекст»).

Уникальность: ~50 шаблонов на категорию × случайные подстановки (имена,
суммы, товары, дни, номера кассы) → десятки тысяч уникальных комбинаций.

Запуск:
    docker run --rm -v "./examples:/work" -w /work llm-backend python _generate_fraud_sample.py

Параметры:
    ROWS_TOTAL — сколько строк сгенерировать (default 100_000). Для быстрого
    локального теста можно запустить с меньшим значением через ENV.
"""
from __future__ import annotations

import os
import random
from datetime import datetime, timedelta

from openpyxl import Workbook

SEED = 20260419
ROWS_TOTAL = int(os.getenv("ROWS_TOTAL", "100000"))

STORES: list[dict] = [
    {"name": "Магазин_Север", "weight": 0.20, "fraud_rate": 0.15},
    {"name": "Магазин_Центр", "weight": 0.20, "fraud_rate": 0.05},
    {"name": "Магазин_Юг", "weight": 0.20, "fraud_rate": 0.00},
    {"name": "Магазин_Запад", "weight": 0.20, "fraud_rate": 0.10},
    {"name": "Магазин_Онлайн", "weight": 0.20, "fraud_rate": 0.00},
]

# Доли остальных категорий (от НЕ-fraud части отзывов магазина).
OTHER_DISTRIBUTION = {
    "service": 0.30,
    "quality": 0.20,
    "positive": 0.35,
    "neutral": 0.15,
}

NAMES_MALE = [
    "Михаил", "Алексей", "Сергей", "Дмитрий", "Андрей", "Игорь", "Николай",
    "Павел", "Константин", "Владимир", "Артём", "Роман", "Виктор", "Евгений",
    "Максим", "Олег", "Юрий", "Денис", "Антон", "Станислав",
]
NAMES_FEMALE = [
    "Елена", "Ольга", "Ирина", "Татьяна", "Наталья", "Анна", "Мария", "Светлана",
    "Екатерина", "Юлия", "Алина", "Вероника", "Кристина", "Валентина", "Ксения",
    "Дарья", "Маргарита", "Полина", "Лидия", "Галина",
]
NAMES_ALL = NAMES_MALE + NAMES_FEMALE

GOODS_GROCERY = [
    "молоко", "хлеб", "йогурт", "сыр", "масло", "колбасу", "куриное филе",
    "творог", "пельмени", "овощи", "фрукты", "консервы", "замороженную пиццу",
    "мороженое", "крупу", "сок",
]
GOODS_NONFOOD = [
    "наушники", "куртку", "термос", "чайник", "пылесос", "монитор", "телефон",
    "клавиатуру", "постельное", "посуду", "косметику", "детский конструктор",
    "зарядку", "фен", "утюг",
]
GOODS_ALL = GOODS_GROCERY + GOODS_NONFOOD

PAYMENT_APPS = ["Тинькофф", "Сбер", "Альфу", "Райф", "ВТБ", "Озон-банк"]
AMOUNTS_SMALL = [350, 480, 690, 820, 1100, 1350, 1580, 1990]
AMOUNTS_MID = [2100, 2470, 2980, 3200, 3700, 4150, 4800, 5200]
AMOUNTS_LARGE = [5800, 6400, 7300, 8100, 9200, 10500, 12800, 14400]

KASSA = [1, 2, 3, 4, 5, 6, 7, 8]

# ───────────────────── ШАБЛОНЫ ───────────────────── #
# Каждый шаблон — callable, возвращает строку с подставленными random-значениями.
# Сгруппированы по схемам внутри категории, чтобы можно было указать
# какую конкретно схему «подкинуть» в каждом магазине.


def rnd(choices):
    return random.choice(choices)


def tpl_fraud_payment_bypass():
    templates = [
        lambda: f"Кассир {rnd(NAMES_ALL)} предложил не пробивать часть товара и перевести {rnd(AMOUNTS_MID)} рублей на {rnd(PAYMENT_APPS)} ему лично. Отказался, но настаивал.",
        lambda: f"На кассе №{rnd(KASSA)} сказали «касса сломана», попросили перевести по номеру телефона — чек отдали рукописный на салфетке.",
        lambda: f"Покупал {rnd(GOODS_NONFOOD)} за {rnd(AMOUNTS_LARGE)}, продавец {rnd(NAMES_ALL)} намекнул что можно оформить без чека со скидкой 20% если скинуть на его карту.",
        lambda: f"Заплатил {rnd(AMOUNTS_MID)} рублей, кассир {rnd(NAMES_ALL)} пробил только половину суммы, предложил остальное положить в конверт — в пользу сотрудников.",
        lambda: f"Вчера вечером на кассе {rnd(NAMES_ALL)} попросил перевести {rnd(AMOUNTS_SMALL)} сверх чека на {rnd(PAYMENT_APPS)} за «срочное оформление возврата».",
        lambda: f"Сотрудник {rnd(NAMES_ALL)} сказал, что если оплатить напрямую ему на Сбер — будет скидка 15%, чек распечатывать не будут.",
        lambda: f"Покупка {rnd(AMOUNTS_LARGE)}, кассир пробил {rnd(AMOUNTS_MID)} и сказал «остальное мне на телефон». Ушла, но ощущение неприятное.",
    ]
    return rnd(templates)()


def tpl_fraud_extortion():
    templates = [
        lambda: f"Охранник {rnd(NAMES_MALE)} на входе требовал {rnd(AMOUNTS_SMALL)} рублей «за ускоренный проход» с крупной тележкой, иначе грозился проверять каждую позицию час.",
        lambda: f"Оператор поддержки в чате намекал, что за {rnd(AMOUNTS_MID)} рублей сверху возврат оформят не за 14 дней, а за один. Это что вообще?",
        lambda: f"Менеджер {rnd(NAMES_ALL)} вымогал оплату сверх чека за «приоритетную обработку заявки» — якобы иначе жалобу потеряют.",
        lambda: f"Курьер потребовал {rnd(AMOUNTS_SMALL)} рублей наличными за подъём на 5 этаж, хотя в договоре этого нет.",
        lambda: f"Завхоз {rnd(NAMES_ALL)} требовал денег за то, чтобы отпустить товар раньше очереди — «иначе ждите своей смены».",
    ]
    return rnd(templates)()


def tpl_fraud_fake_discount():
    templates = [
        lambda: f"Продавец {rnd(NAMES_FEMALE)} предложила купить {rnd(GOODS_NONFOOD)} за {rnd(AMOUNTS_MID)} вместо {rnd(AMOUNTS_LARGE)}, но «только через мою карту и мой чек».",
        lambda: f"Сотрудник {rnd(NAMES_ALL)} сказал, что у него есть «персональная скидка для своих» на {rnd(GOODS_NONFOOD)} — надо только перевести ему напрямую.",
        lambda: f"Мерчандайзер настаивал выкупить уценённый товар через его личную карту, якобы по специальному тарифу. Подозрительно.",
        lambda: f"Курьер оформил «скидку» {rnd(AMOUNTS_SMALL)} рублей и попросил часть суммы налом на руки, без отражения в системе.",
    ]
    return rnd(templates)()


def tpl_fraud_product_swap():
    templates = [
        lambda: f"Продавец {rnd(NAMES_ALL)} прямо предложил купить витринный {rnd(GOODS_NONFOOD)} со следами использования как новый — со скидкой, но за нал в обход кассы.",
        lambda: f"Пришёл {rnd(GOODS_NONFOOD)} — упаковка вскрыта, царапины. Сотрудник пункта выдачи настаивал, что «так и было», и предлагал доплату за обмен мимо системы.",
        lambda: f"Кладовщик {rnd(NAMES_MALE)} предложил купить восстановленную технику по цене новой — с «его гарантией» и без чека.",
    ]
    return rnd(templates)()


def tpl_fraud_return_scam():
    templates = [
        lambda: f"Пришёл на возврат {rnd(GOODS_NONFOOD)} за {rnd(AMOUNTS_LARGE)}, сотрудник {rnd(NAMES_ALL)} оформил «частичный возврат» — деньги пришли меньше, разницу предложил отдать налом «в следующий раз».",
        lambda: f"Заметил что при возврате бланк заполнен карандашом. Кассир {rnd(NAMES_ALL)} сказал «так принято», но цифры потом исправили — вернули меньше заявленного.",
        lambda: f"Возврат оформили не полностью, обещали разницу отправить СБП — до сих пор ничего не пришло, сотрудник не отвечает.",
    ]
    return rnd(templates)()


def tpl_service_rude():
    templates = [
        lambda: f"Продавец {rnd(NAMES_ALL)} разговаривал сквозь зубы, на вопрос про {rnd(GOODS_ALL)} ответил «ищите сами, я не справочная».",
        lambda: f"Кассир {rnd(NAMES_ALL)} хамила женщине передо мной — и потом мне. День испорчен с самого начала.",
        lambda: f"На горячей линии {rnd(NAMES_FEMALE)} прервала меня на полуслове и повесила трубку, пришлось звонить заново.",
        lambda: f"Консультант {rnd(NAMES_ALL)} закатывала глаза на каждый вопрос. Такое ощущение что я её раздражаю.",
        lambda: f"Охранник на входе хамовато спросил «куда прёмся» и оскорбил пакет с покупками.",
        lambda: f"Оператор в чате скопировал три раза одну и ту же отписку и перестал отвечать.",
        lambda: f"Заведующий {rnd(NAMES_ALL)} общался как будто одолжение делает, говорил свысока на просьбу посмотреть ценник.",
    ]
    return rnd(templates)()


def tpl_service_slow():
    templates = [
        lambda: f"Стоял в очереди {random.randint(25, 55)} минут, работала только касса №{rnd(KASSA)} из пяти.",
        lambda: f"Курьер опоздал на {random.randint(2, 5)} часа, не перезванивал, заказ приехал уже остывший.",
        lambda: f"Консультант искал ответ {random.randint(15, 40)} минут, в итоге отправил к другому сотруднику.",
        lambda: f"Обещанная доставка «в течение двух часов» превратилась в полдня. Никаких уведомлений.",
        lambda: f"В пункте выдачи один сотрудник на 30 человек — застрял на 40 минут, заказ лежал прямо рядом.",
        lambda: f"Поддержка отвечает по 6-8 часов, проблему не решают с первого обращения, отсылают на FAQ.",
    ]
    return rnd(templates)()


def tpl_service_dirty():
    templates = [
        lambda: f"В зале грязные полы, корзины липкие, на кассе №{rnd(KASSA)} сломан дисплей.",
        lambda: f"Курьер принёс еду в помятой коробке, пакет грязный. Еду открывать было неприятно.",
        lambda: f"В раздевалке примерочной нет крючков, приходится класть вещи прямо на пол — не очень приятно.",
        lambda: f"Туалет закрыт на ремонт уже третий месяц, в зале душно, вентиляция не работает.",
        lambda: f"Молоко на полке было с заляпанными этикетками, видно что пакеты кто-то переставлял в лужу.",
    ]
    return rnd(templates)()


def tpl_service_no_response():
    templates = [
        lambda: f"Написал в чат три дня назад про возврат — ни одного ответа, только автоответчик.",
        lambda: f"Оставил претензию через сайт, ID №{random.randint(100000, 999999)}, прошла неделя — молчание.",
        lambda: f"Телефон поддержки играет мелодию {random.randint(20, 50)} минут и обрывается. Это что, такая проверка терпения?",
        lambda: f"Заполнил форму обратной связи дважды, пришло два одинаковых автоответа «спасибо, мы разбираемся».",
    ]
    return rnd(templates)()


def tpl_quality_spoiled():
    templates = [
        lambda: f"Купил {rnd(GOODS_GROCERY)} — срок годности до {random.randint(20, 30)}.04.26, а уже прокисло. Выбросил.",
        lambda: f"Хлеб свежий снаружи, разрезал — плесень внутри. Как можно такое продавать.",
        lambda: f"Молоко свернулось через день хранения в холодильнике, хотя срок ещё неделю.",
        lambda: f"Колбаса пахла не очень уже при вскрытии вакуума. Скорее всего переупаковали просрочку.",
        lambda: f"Йогурт — на дне крышки плесневые пятна, при том что срок годности ещё неделю.",
    ]
    return rnd(templates)()


def tpl_quality_broken():
    templates = [
        lambda: f"Заказал {rnd(GOODS_NONFOOD)} — пришёл с трещиной на корпусе. Упаковка целая, значит брак изначальный.",
        lambda: f"Купил {rnd(GOODS_NONFOOD)} за {rnd(AMOUNTS_LARGE)} — оказался бракованным на третий день. Возврат оформили со скрипом.",
        lambda: f"{rnd(GOODS_NONFOOD).capitalize()} шумит как трактор. Подозреваю восстановленный экземпляр под видом нового.",
        lambda: f"Пришёл не тот размер — заказывал M, прислали XL. Переделывать отказались, только возврат.",
        lambda: f"Наушники хрипят на правом канале с первого включения. Гарантия есть, но менять дольше месяца.",
    ]
    return rnd(templates)()


def tpl_quality_mismatch():
    templates = [
        lambda: f"На сайте {rnd(GOODS_NONFOOD)} был синий, привезли чёрный. Пересорт, но менять отказываются — «фото не договор».",
        lambda: f"Заказывал одну модель {rnd(GOODS_NONFOOD)}, в коробке лежит другая. Серийник не совпадает ни с чем.",
        lambda: f"Куртка с дыркой на спине, в описании ни слова, фото сделаны идеальные.",
        lambda: f"Детский конструктор из описания на 200 деталей — внутри комплект на 120, остальные «уточняйте».",
    ]
    return rnd(templates)()


def tpl_positive():
    templates = [
        lambda: f"Отличный магазин, {rnd(GOODS_ALL)} всегда свежие, кассир {rnd(NAMES_ALL)} даже улыбается — редкость.",
        lambda: f"Спасибо продавцу {rnd(NAMES_FEMALE)} — помогла подобрать {rnd(GOODS_NONFOOD)}, рассказала плюсы-минусы, без впаривания.",
        lambda: f"Курьер привёз точно в срок, вежливый, сам предложил перенести на час раньше — согласовал по телефону.",
        lambda: f"Акция реальная, ценник в зале совпал с чеком. После истории «Пятёрочки» — большая редкость.",
        lambda: f"Очень понравилось: заказ оформил в три клика, доставили на следующий день, упаковка аккуратная.",
        lambda: f"Консультант {rnd(NAMES_ALL)} помог выбрать {rnd(GOODS_NONFOOD)}, объяснил плюсы-минусы — купил сразу.",
        lambda: f"Возврат оформили без вопросов за 10 минут, деньги пришли через день. Молодцы.",
        lambda: f"Приложение удобное, Apple Pay работает, бонусы начисляются сразу и их можно тратить.",
        lambda: f"Отличное качество {rnd(GOODS_GROCERY)}, беру здесь уже третий месяц. Не разочаровывает.",
        lambda: f"Поддержка ответила за минуту, решила проблему сразу — без десятка переадресаций.",
        lambda: f"Продавец {rnd(NAMES_ALL)} не побоялся сказать, что {rnd(GOODS_NONFOOD)} в соседнем магазине дешевле. Честность подкупает.",
        lambda: f"Ассортимент радует, нашёл редкий {rnd(GOODS_NONFOOD)} которого у конкурентов нет.",
        lambda: f"В зале чисто, очередей нет, персонал приветливый. Хочется возвращаться.",
        lambda: f"Курьер позвонил за 15 минут, вежливо уточнил этаж и код домофона. Всё по-человечески.",
    ]
    return rnd(templates)()


def tpl_neutral():
    templates = [
        lambda: f"Нормально, без восторга и без претензий. Купил что хотел, ушёл.",
        lambda: f"В принципе ок, но {rnd(GOODS_GROCERY)} могли бы и посвежее привезти. Не критично.",
        lambda: f"Магазин как магазин. Ни плюсов особых, ни минусов.",
        lambda: f"Цены выше среднего, но и выбор шире — на любителя.",
        lambda: f"Иногда работают акции, иногда нет — зависит от смены. Неровно.",
        lambda: f"Приложение подвисает на оплате, но в итоге проходит. Неудобно, но не критично.",
        lambda: f"Доставка пришла позже обещанного на 40 минут — не смертельно, но неприятно.",
        lambda: f"Персонал разный: одни приветливы, другие сквозь зубы. Рулетка.",
        lambda: f"Парковка маленькая, придётся постоять. На машине лучше в будни.",
        lambda: f"Средне, ожидал лучшего за такие деньги.",
    ]
    return rnd(templates)()


# ───────── «Характеры» магазинов ─────────
# fraud_mix — какие схемы мошенничества встречаются. Разные профили для
# разных магазинов — чтобы групповой анализ вытягивал разные паттерны.

STORE_FRAUD_MIX = {
    "Магазин_Север": [  # кассовая дисциплина + вымогательство
        (tpl_fraud_payment_bypass, 0.5),
        (tpl_fraud_extortion, 0.3),
        (tpl_fraud_return_scam, 0.2),
    ],
    "Магазин_Запад": [  # курьерские схемы + подмены
        (tpl_fraud_product_swap, 0.4),
        (tpl_fraud_fake_discount, 0.4),
        (tpl_fraud_payment_bypass, 0.2),
    ],
    "Магазин_Центр": [  # единичные случаи, ассорти
        (tpl_fraud_payment_bypass, 0.4),
        (tpl_fraud_fake_discount, 0.3),
        (tpl_fraud_extortion, 0.3),
    ],
    # Юг и Онлайн — без fraud.
}

# service/quality — общие шаблоны для всех магазинов
SERVICE_POOL = [
    (tpl_service_rude, 0.35),
    (tpl_service_slow, 0.30),
    (tpl_service_dirty, 0.20),
    (tpl_service_no_response, 0.15),
]
QUALITY_POOL = [
    (tpl_quality_spoiled, 0.35),
    (tpl_quality_broken, 0.40),
    (tpl_quality_mismatch, 0.25),
]


def pick_weighted(pool: list[tuple]) -> str:
    """Выбор функции-шаблона по весам, вызов."""
    templates, weights = zip(*pool)
    return random.choices(templates, weights=weights, k=1)[0]()


def category_for_store(store: dict) -> tuple[str, callable]:
    """Возвращает (имя категории, функция-шаблон) для одного отзыва магазина."""
    roll = random.random()
    if roll < store["fraud_rate"] and store["name"] in STORE_FRAUD_MIX:
        return "fraud", pick_weighted_factory(STORE_FRAUD_MIX[store["name"]])
    # оставшаяся вероятностная масса — распределение между service/quality/positive/neutral
    residual = random.random()
    acc = 0.0
    for cat, share in OTHER_DISTRIBUTION.items():
        acc += share
        if residual < acc:
            if cat == "service":
                return cat, lambda: pick_weighted(SERVICE_POOL)
            if cat == "quality":
                return cat, lambda: pick_weighted(QUALITY_POOL)
            if cat == "positive":
                return cat, tpl_positive
            return cat, tpl_neutral
    return "neutral", tpl_neutral


def pick_weighted_factory(pool: list[tuple]):
    def inner():
        return pick_weighted(pool)
    return inner


# ───────── Основная генерация ─────────

def generate_shift_label(store_short: str, day: datetime, slots_per_day: int) -> str:
    """Формат смены: «Север_2026-03-15_утро». Число слотов — адаптивно от объёма файла."""
    if slots_per_day >= 3:
        slot = rnd(["утро", "день", "вечер"])
    elif slots_per_day == 2:
        slot = rnd(["утро", "вечер"])
    else:
        slot = "смена"
    return f"{store_short}_{day.strftime('%Y-%m-%d')}_{slot}"


STORE_SHORT = {
    "Магазин_Север": "Север",
    "Магазин_Центр": "Центр",
    "Магазин_Юг": "Юг",
    "Магазин_Запад": "Запад",
    "Магазин_Онлайн": "Онлайн",
}


def rating_for_category(cat: str) -> int:
    if cat == "positive":
        return rnd([4, 5, 5, 5])
    if cat == "neutral":
        return rnd([3, 3, 4])
    if cat == "fraud":
        return rnd([1, 1, 2, 3])
    if cat == "service":
        return rnd([1, 1, 2, 2])
    if cat == "quality":
        return rnd([1, 2, 2])
    return 3


def main() -> None:
    random.seed(SEED)

    # Адаптивная сетка смен: цель — средняя группа ~40-60 отзывов/смену.
    # При меньших файлах сокращаем период и число слотов, иначе смен больше строк
    # и «групповая» группировка вырождается до одной строки в группе.
    target_group_size = 45
    total_shifts_target = max(5, ROWS_TOTAL // target_group_size)
    total_shifts_per_store = max(1, total_shifts_target // len(STORES))
    if total_shifts_per_store >= 90:
        slots_per_day, period_days = 3, min(150, total_shifts_per_store // 3)
    elif total_shifts_per_store >= 30:
        slots_per_day, period_days = 2, total_shifts_per_store // 2
    else:
        slots_per_day, period_days = 1, total_shifts_per_store
    period_days = max(1, period_days)

    start = datetime(2026, 1, 1)
    end = start + timedelta(days=period_days, hours=23, minutes=59)
    total_seconds = int((end - start).total_seconds())

    print(f"Генерация {ROWS_TOTAL} отзывов…")
    print(f"Период: {period_days} дн × {slots_per_day} слот(а) × 5 магазинов ≈ {period_days * slots_per_day * 5} смен")

    # Предрассчитываем сколько строк каждому магазину (детерминированно по весам).
    quotas: dict[str, int] = {}
    remaining = ROWS_TOTAL
    for idx, s in enumerate(STORES):
        if idx == len(STORES) - 1:
            quotas[s["name"]] = remaining
        else:
            q = int(ROWS_TOTAL * s["weight"])
            quotas[s["name"]] = q
            remaining -= q
    print(f"Квоты по магазинам: {quotas}")

    rows: list[tuple] = []
    row_id = 0
    category_stats: dict[str, int] = {"fraud": 0, "service": 0, "quality": 0, "positive": 0, "neutral": 0}

    for store in STORES:
        store_name = store["name"]
        store_short = STORE_SHORT[store_name]
        for _ in range(quotas[store_name]):
            row_id += 1
            cat, template_fn = category_for_store(store)
            text = template_fn()
            rating = rating_for_category(cat)
            seconds_offset = random.randint(0, total_seconds)
            moment = start + timedelta(seconds=seconds_offset)
            shift = generate_shift_label(store_short, moment, slots_per_day)
            rows.append(
                (
                    row_id,
                    store_name,
                    shift,
                    text,
                    rating,
                    moment.strftime("%Y-%m-%d %H:%M:%S"),
                )
            )
            category_stats[cat] += 1

    random.shuffle(rows)  # перемешиваем, чтобы не было 20k подряд из одного магазина
    # После shuffle — id пересоздадим по порядку для аккуратности.
    rows = [(idx + 1, r[1], r[2], r[3], r[4], r[5]) for idx, r in enumerate(rows)]

    # Считаем распределение по магазинам для лога.
    from collections import Counter
    by_store = Counter(r[1] for r in rows)
    shifts = Counter(r[2] for r in rows)
    print("Итоговое распределение категорий:")
    for k, v in category_stats.items():
        print(f"  {k}: {v} ({v / ROWS_TOTAL * 100:.1f}%)")
    print(f"Магазинов: {len(by_store)}, смен: {len(shifts)}, средняя группа по смене: {ROWS_TOTAL / len(shifts):.1f}")

    print("Запись xlsx…")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Reviews")
    ws.append(["id", "магазин", "смена", "отзыв", "оценка", "дата"])
    for row in rows:
        ws.append(row)

    # AltSheet для совместимости тестов инспекта (ранее был в маленьком sample).
    alt = wb.create_sheet("AltSheet")
    alt.append(["comment"])
    alt.append(["Тестовая строка на втором листе"])

    out_name = f"Мошенничество_{ROWS_TOTAL}_отзывов.xlsx"
    wb.save(out_name)
    print(f"Сохранено: {out_name}")


if __name__ == "__main__":
    main()
